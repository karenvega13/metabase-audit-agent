"""
reverse_propagate.py
Feedback loop inverso: propaga las decisiones de los JSONs de resultados
(incluidas las clasificadas por auto_classify.py) de vuelta al tracker Excel,
actualizando las filas que aún están en pending_validation.

Uso:
    python reverse_propagate.py --inspect     # stats sin tocar nada
    python reverse_propagate.py --dry-run     # muestra qué actualizaría
    python reverse_propagate.py --diagnose    # muestra filas sin match lado a lado
    python reverse_propagate.py               # ejecuta la propagación real

Qué hace:
    1. Lee todos los JSONs de resultados (lote_*_resultados.json)
    2. Construye un índice de decisiones: {card_id: [{status, descripcion, notas, fuente}]}
    3. Lee la hoja "🔍 Todos los hallazgos" del tracker Excel
    4. Para cada fila con Validation Status = "pending_validation":
       a. Busca el card_id en el índice JSON
       b. Hace matching de descripción (mismo algoritmo que propagate_reviews.py)
       c. Si matchea, escribe el status/notas/validated_by/validated_date en Excel
    5. Hace backup del Excel antes de escribir (solo en ejecución real)

Reglas de seguridad:
    - NUNCA sobreescribe una fila con status ya decidido (confirmed_error/false_positive/intentional)
    - Solo toca filas donde Validation Status = "pending_validation" o vacío
    - El backup se crea automáticamente antes de cualquier escritura real
"""

import sys
import os
import re
import json
import shutil
import argparse
import openpyxl
from pathlib import Path
from datetime import date
from collections import defaultdict

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ============================================================
# CONFIG
# ============================================================

_ROOT        = Path(__file__).resolve().parent.parent.parent.parent
TRACKER_FILE = str(_ROOT / "data" / "raw" / "tracker_auditoria_metabase.xlsx")
RESULTS_DIR  = str(_ROOT / "data" / "processed" / "resultados")
MAIN_SHEET   = "\U0001f50d Todos los hallazgos"
HEADER_ROW   = 2
DATA_ROW_FROM = 3

# Columnas (mismas que propagate_reviews.py)
COL_CARD_ID      = "Card ID"
COL_STATUS       = "Validation Status"
COL_DESCRIPCION  = "Descripcion Hallazgo"   # alias por si hay variante sin tilde
COL_DESCRIPCION2 = "Descripción Hallazgo"   # versión con tilde
COL_NOTAS        = "Notas Verificacion"
COL_NOTAS2       = "Notas Verificación"
COL_VALIDATED_BY = "Validated By"
COL_VALIDATED_DT = "Validated Date"

DECIDED_STATUSES = {"confirmed_error", "false_positive", "intentional"}
PENDING_VALUES   = {"pending_validation", "pending", "", None}

TODAY = date.today().isoformat()


# ============================================================
# MATCHING (reutilizado de propagate_reviews.py)
# ============================================================

def _token_overlap(a: str, b: str) -> float:
    ta = set(re.findall(r"\w+", a.lower()))
    tb = set(re.findall(r"\w+", b.lower()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def match_description(json_desc: str, tracker_desc: str,
                      fuzzy_threshold: float = 0.60) -> tuple[bool, str]:
    """
    Devuelve (matched, method). Misma lógica que propagate_reviews.py:
      1. Exacto en primeros 80 chars
      2. Fragmento de 40 chars
      3. Token overlap >= fuzzy_threshold
    """
    if not tracker_desc or not json_desc:
        return False, ""

    t_short = tracker_desc[:80].strip().lower()
    j_short = json_desc[:80].strip().lower()

    if t_short == j_short:
        return True, "exact"

    if len(t_short) >= 40 and t_short[:40] in j_short:
        return True, "exact"
    if len(j_short) >= 40 and j_short[:40] in t_short:
        return True, "exact"

    score = _token_overlap(tracker_desc, json_desc)
    if score >= fuzzy_threshold:
        return True, f"fuzzy_{score:.2f}"

    return False, ""


# ============================================================
# CARGAR DECISIONES DESDE JSONs
# ============================================================

def load_json_decisions(results_dir: str) -> dict:
    """
    Lee todos los lote_*_resultados.json y construye:
    {card_id: [{"status", "descripcion", "notas", "validated_by", "validated_date"}, ...]}

    Solo incluye hallazgos con status decidido (confirmed_error / false_positive / intentional).
    """
    decisions = defaultdict(list)
    json_files = sorted(Path(results_dir).glob("lote_*_resultados.json"))

    for jf in json_files:
        try:
            data = json.load(open(jf, encoding="utf-8"))
        except Exception:
            continue

        cards = data if isinstance(data, list) else data.get("cards", [])
        for card in cards:
            if not isinstance(card, dict):
                continue
            cid = card.get("card_id")
            if not cid:
                continue
            for h in card.get("hallazgos", []):
                status = h.get("human_review_status", "")
                if status not in DECIDED_STATUSES:
                    continue
                decisions[cid].append({
                    "status":       status,
                    "descripcion":  h.get("descripcion", ""),
                    "notas":        h.get("human_review_notes") or h.get("validated_by") or "",
                    "validated_by": h.get("validated_by") or "auto_classify.py / S56",
                    "validated_date": h.get("validated_date") or TODAY,
                })

    return dict(decisions)


# ============================================================
# CARGAR TRACKER EXCEL
# ============================================================

def load_tracker_pending(filepath: str) -> tuple[list, dict, dict]:
    """
    Lee el tracker y devuelve:
      - rows: lista de dicts con todos los datos de cada fila (incluyendo row_number)
      - headers: {nombre_col: indice_0based}
      - stats: resumen
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if MAIN_SHEET not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Hoja '{MAIN_SHEET}' no encontrada. Disponibles: {available}")

    ws = wb[MAIN_SHEET]

    # Leer headers (fila HEADER_ROW)
    headers = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for col_idx, h in enumerate(row, 1):
            if h is not None:
                headers[str(h).strip()] = col_idx - 1  # 0-based
        break

    # Resolver alias de columnas con/sin tilde
    def resolve(primary, fallback):
        if primary in headers:
            return primary
        if fallback in headers:
            return fallback
        return None

    col_card_id   = resolve(COL_CARD_ID, COL_CARD_ID)
    col_status    = resolve(COL_STATUS, COL_STATUS)
    col_desc      = resolve(COL_DESCRIPCION2, COL_DESCRIPCION)
    col_notas     = resolve(COL_NOTAS2, COL_NOTAS)
    col_val_by    = resolve(COL_VALIDATED_BY, COL_VALIDATED_BY)
    col_val_dt    = resolve(COL_VALIDATED_DT, COL_VALIDATED_DT)

    required_cols = [col_card_id, col_status, col_desc]
    if any(c is None for c in required_cols):
        missing = [n for n, c in [("Card ID", col_card_id), ("Validation Status", col_status), ("Descripcion", col_desc)] if c is None]
        raise ValueError(f"Columnas requeridas no encontradas: {missing}. Headers: {list(headers.keys())}")

    idx_card_id = headers[col_card_id]
    idx_status  = headers[col_status]
    idx_desc    = headers[col_desc]
    idx_notas   = headers.get(col_notas, -1) if col_notas else -1
    idx_val_by  = headers.get(col_val_by, -1) if col_val_by else -1
    idx_val_dt  = headers.get(col_val_dt, -1) if col_val_dt else -1

    rows = []
    total = pending_count = decided_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=DATA_ROW_FROM, values_only=True), start=DATA_ROW_FROM):
        raw_id = row[idx_card_id] if idx_card_id < len(row) else None
        if raw_id is None:
            continue
        try:
            card_id = int(str(raw_id).strip())
        except (ValueError, TypeError):
            continue

        total += 1
        raw_status = row[idx_status] if idx_status < len(row) else None
        status = str(raw_status).strip().lower() if raw_status else ""

        if status in DECIDED_STATUSES:
            decided_count += 1
            continue  # no tocar filas ya decididas

        # Solo procesar pending
        pending_count += 1
        desc = str(row[idx_desc]) if idx_desc < len(row) and row[idx_desc] else ""
        notas = str(row[idx_notas]) if idx_notas >= 0 and idx_notas < len(row) and row[idx_notas] else ""
        val_by = str(row[idx_val_by]) if idx_val_by >= 0 and idx_val_by < len(row) and row[idx_val_by] else ""
        val_dt = str(row[idx_val_dt]) if idx_val_dt >= 0 and idx_val_dt < len(row) and row[idx_val_dt] else ""

        rows.append({
            "row_num":      row_num,
            "card_id":      card_id,
            "status":       status,
            "descripcion":  desc,
            "notas":        notas,
            "validated_by": val_by,
            "validated_date": val_dt,
        })

    col_indices = {
        "status":       idx_status,
        "notas":        idx_notas,
        "validated_by": idx_val_by,
        "validated_date": idx_val_dt,
    }

    stats = {
        "total":     total,
        "decided":   decided_count,
        "pending":   pending_count,
    }

    return rows, col_indices, stats


# ============================================================
# MATCHING: JSON decisions → Excel pending rows
# ============================================================

def match_pending_rows(pending_rows: list, json_decisions: dict) -> list:
    """
    Para cada fila pending del Excel, busca la decision correspondiente en el JSON.
    Devuelve lista de dicts con el match (o sin match):
      {row_num, card_id, matched, method, json_status, json_notas, json_validated_by, json_validated_date, tracker_desc, json_desc}
    """
    results = []

    for row in pending_rows:
        cid  = row["card_id"]
        tdesc = row["descripcion"]

        if cid not in json_decisions:
            results.append({**row, "matched": False, "method": "no_json", "json_status": None})
            continue

        json_hallazgos = json_decisions[cid]

        # Intentar match por descripcion
        best_match = None
        best_method = ""
        used = set()

        for j_idx, jh in enumerate(json_hallazgos):
            if j_idx in used:
                continue
            ok, method = match_description(jh["descripcion"], tdesc)
            if ok:
                best_match  = jh
                best_method = method
                used.add(j_idx)
                break

        # Force-single: si no hubo match pero hay 1 JSON hallazgo y 1 Excel row para esta card
        if not best_match and len(json_hallazgos) == 1:
            # Contar cuántas filas pending del Excel tienen este card_id
            same_card_rows = [r for r in pending_rows if r["card_id"] == cid]
            if len(same_card_rows) == 1:
                best_match  = json_hallazgos[0]
                best_method = "force_single"

        if best_match:
            results.append({
                **row,
                "matched":          True,
                "method":           best_method,
                "json_status":      best_match["status"],
                "json_notas":       best_match["notas"],
                "json_validated_by": best_match["validated_by"],
                "json_validated_date": best_match["validated_date"],
                "json_desc":        best_match["descripcion"],
            })
        else:
            results.append({
                **row,
                "matched":   False,
                "method":    "no_match",
                "json_status": None,
                "json_desc": json_hallazgos[0]["descripcion"] if json_hallazgos else "",
            })

    return results


# ============================================================
# ESCRIBIR EN EXCEL
# ============================================================

def write_to_excel(filepath: str, matches: list, col_indices: dict) -> int:
    """
    Escribe las decisiones matcheadas en el Excel.
    Devuelve el número de filas actualizadas.
    """
    # Backup antes de escribir
    backup_path = filepath.replace(".xlsx", f"_backup_S56.xlsx")
    if not os.path.exists(backup_path):
        shutil.copy2(filepath, backup_path)
        print(f"   Backup creado: {backup_path}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb[MAIN_SHEET]

    idx_status  = col_indices["status"]      # 0-based
    idx_notas   = col_indices["notas"]
    idx_val_by  = col_indices["validated_by"]
    idx_val_dt  = col_indices["validated_date"]

    # Convertir de 0-based a 1-based para openpyxl
    def col1(idx):
        return idx + 1

    updated = 0
    for m in matches:
        if not m["matched"]:
            continue
        row_num = m["row_num"]
        ws.cell(row=row_num, column=col1(idx_status)).value = m["json_status"]
        if idx_notas >= 0:
            ws.cell(row=row_num, column=col1(idx_notas)).value = m["json_notas"] or ""
        if idx_val_by >= 0:
            ws.cell(row=row_num, column=col1(idx_val_by)).value = m["json_validated_by"] or "auto_classify.py"
        if idx_val_dt >= 0:
            ws.cell(row=row_num, column=col1(idx_val_dt)).value = m["json_validated_date"] or TODAY
        updated += 1

    wb.save(filepath)
    return updated


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Feedback loop inverso: JSON decisions → Excel pending rows"
    )
    parser.add_argument("--inspect",  action="store_true",
                        help="Solo stats: no modifica nada")
    parser.add_argument("--dry-run",  action="store_true",
                        help="Muestra qué actualizaria sin escribir")
    parser.add_argument("--diagnose", action="store_true",
                        help="Muestra filas sin match lado a lado")
    args = parser.parse_args()

    mode = "INSPECT" if args.inspect else "DRY-RUN" if args.dry_run else "REAL"
    print("=" * 62)
    print("  Reverse Propagate — JSONs -> Excel pending rows")
    print(f"  Tracker: {TRACKER_FILE}")
    print(f"  Modo: {mode}")
    print("=" * 62)

    # 1. Cargar decisiones de JSONs
    print("\n  Cargando decisiones de JSONs...")
    json_decisions = load_json_decisions(RESULTS_DIR)
    total_json_hallazgos = sum(len(v) for v in json_decisions.values())
    print(f"  Cards con decision en JSON:   {len(json_decisions):>5}")
    print(f"  Hallazgos decididos en JSON:  {total_json_hallazgos:>5}")

    if not json_decisions:
        print("  Sin decisiones en JSONs. Corre auto_classify.py primero.")
        return

    if args.inspect:
        # Cargar tambien el tracker para dar stats completas
        try:
            _, _, stats = load_tracker_pending(TRACKER_FILE)
            print(f"\n  Tracker Excel:")
            print(f"    Total filas:         {stats['total']:>5}")
            print(f"    Con decision:        {stats['decided']:>5}")
            print(f"    Pending (candidatas): {stats['pending']:>5}")
        except Exception as e:
            print(f"  No se pudo leer tracker: {e}")
        print("\n  Modo --inspect: no se modifico nada.")
        return

    # 2. Cargar filas pending del tracker
    print("\n  Leyendo tracker Excel (filas pending)...")
    try:
        pending_rows, col_indices, stats = load_tracker_pending(TRACKER_FILE)
    except Exception as e:
        print(f"  Error al leer tracker: {e}")
        return

    print(f"  Total filas:          {stats['total']:>5}")
    print(f"  Ya decididas (skip):  {stats['decided']:>5}")
    print(f"  Pending (candidatas): {stats['pending']:>5}")

    if not pending_rows:
        print("\n  No hay filas pending en el tracker. Nada que sincronizar.")
        return

    # 3. Matching
    print("\n  Matcheando descripciones JSON <-> Excel...")
    matches = match_pending_rows(pending_rows, json_decisions)

    matched     = [m for m in matches if m["matched"]]
    no_match    = [m for m in matches if not m["matched"]]
    no_json     = [m for m in no_match if m.get("method") == "no_json"]
    sin_match   = [m for m in no_match if m.get("method") != "no_json"]

    by_method   = {}
    for m in matched:
        by_method[m["method"]] = by_method.get(m["method"], 0) + 1

    print(f"\n  Resultados del matching:")
    print(f"    Matcheadas (listas para sync): {len(matched):>5}")
    for method, cnt in sorted(by_method.items()):
        print(f"      {method}: {cnt}")
    print(f"    Sin JSON para este card_id:    {len(no_json):>5}")
    print(f"    JSON existe pero sin match:    {len(sin_match):>5}")

    if args.diagnose and sin_match:
        print(f"\n  --- DIAGNOSE: {len(sin_match)} filas sin match ---")
        for m in sin_match[:30]:
            print(f"\n  Card #{m['card_id']}:")
            print(f"    Excel desc: {m['descripcion'][:120]}")
            print(f"    JSON  desc: {m.get('json_desc','')[:120]}")

    if args.dry_run:
        print(f"\n  DRY-RUN: se actualizarian {len(matched)} filas en el Excel.")
        print("  Muestra de lo que se escribiria:")
        for m in matched[:5]:
            print(f"    fila {m['row_num']:>4} | Card #{m['card_id']:>5} | {m['json_status']:<20} | {m['method']}")
        print("\n  Modo --dry-run: no se escribio nada.")
        return

    if len(matched) == 0:
        print("\n  Nada para sincronizar (0 matches).")
        return

    # 4. Escribir al Excel
    print(f"\n  Escribiendo {len(matched)} filas en el tracker Excel...")
    try:
        updated = write_to_excel(TRACKER_FILE, matches, col_indices)
    except Exception as e:
        print(f"  ERROR al escribir en Excel: {e}")
        return

    print(f"\n{'=' * 62}")
    print(f"  COMPLETADO")
    print(f"{'=' * 62}")
    print(f"  Filas actualizadas en Excel: {updated:>5}")
    print(f"  Filas sin match (pendientes): {len(sin_match):>4}")
    print(f"  Filas sin JSON:               {len(no_json):>4}")
    print()
    print("  Siguiente paso:")
    if sin_match:
        print(f"    - Revisar {len(sin_match)} filas sin match con --diagnose")
    print("    - Abrir tracker Excel y verificar columnas actualizadas")


if __name__ == "__main__":
    main()
