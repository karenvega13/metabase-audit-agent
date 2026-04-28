"""
propagate_reviews.py
Feedback loop: propaga las decisiones de revisión humana del tracker Excel
hacia los JSONs de resultados del agente auditor.

Uso:
    python propagate_reviews.py --inspect           # muestra stats sin tocar nada
    python propagate_reviews.py --dry-run           # muestra qué actualizaría
    python propagate_reviews.py                     # ejecuta la propagación real

Qué hace:
    1. Lee la hoja "🔍 Todos los hallazgos" del tracker Excel
    2. Extrae todas las filas con Validation Status ≠ 'pending_validation'
    3. Busca cada card_id en los JSONs de resultados
    4. En cada hallazgo de la card, inyecta el status si la descripción coincide
    5. Deriva un status a nivel card (el más severo de sus hallazgos)
    6. Guarda los JSONs actualizados

Campos que agrega a cada CARD en el JSON:
    "human_review_status":  "confirmed_error" | "false_positive" | "intentional" | "mixed"
    "human_review_summary": resumen de cuántos hallazgos de cada tipo

Campos que agrega a cada HALLAZGO en el JSON:
    "human_review_status":  "confirmed_error" | "false_positive" | "intentional" | "pending_validation"
    "human_review_notes":   texto libre del campo Notas Verificación
    "validated_by":         nombre del revisor
    "validated_date":       fecha de validación

Columnas del tracker usadas (hoja "🔍 Todos los hallazgos", fila de header = 2):
    Card ID | Nombre Card | Vistas | Lote | Patrón | Descripción Patrón |
    Severidad | Descripción Hallazgo | Fix Propuesto | Validation Status |
    Notas Verificación | Owner | Validated By | Validated Date
"""

import sys
import os
import re
import json
import argparse
import openpyxl
from pathlib import Path
from datetime import datetime
from collections import defaultdict

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Token overlap para matching fuzzy
def _token_overlap(a: str, b: str) -> float:
    ta = set(re.findall(r"\w+", a.lower()))
    tb = set(re.findall(r"\w+", b.lower()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)

# ============================================================
# CONFIG
# ============================================================

_ROOT         = Path(__file__).resolve().parent.parent.parent.parent
TRACKER_FILE  = str(_ROOT / "data" / "raw" / "tracker_auditoria_metabase.xlsx")
OUTPUT_DIR    = str(_ROOT / "data" / "processed" / "resultados")
MAIN_SHEET    = "🔍 Todos los hallazgos"
HEADER_ROW    = 2      # fila 1 = título, fila 2 = headers reales
DATA_ROW_FROM = 3      # los datos empiezan en fila 3

# Columnas esperadas en la hoja (nombres exactos del tracker)
COL_CARD_ID      = "Card ID"
COL_STATUS       = "Validation Status"
COL_NOTAS        = "Notas Verificación"
COL_DESCRIPCION  = "Descripción Hallazgo"
COL_VALIDATED_BY = "Validated By"
COL_VALIDATED_DT = "Validated Date"

# Valores de status que ya tienen decisión (los demás se saltan)
DECIDED_STATUSES = {"confirmed_error", "false_positive", "intentional"}

# Prioridad para derivar status a nivel card (mayor número = más severo)
STATUS_PRIORITY = {
    "false_positive": 1,
    "intentional":    2,
    "confirmed_error": 3,
}


# ============================================================
# LEER TRACKER
# ============================================================

def load_tracker(filepath: str) -> tuple[dict, dict]:
    """
    Lee el tracker Excel y devuelve:
      - decisions: {card_id (int): [{"status", "descripcion", "notas", "validated_by", "validated_date"}, ...]}
      - stats:     {"total_rows", "decididas", "pending", "por_status": {...}}

    Solo incluye filas con status en DECIDED_STATUSES.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Tracker no encontrado: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if MAIN_SHEET not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Hoja '{MAIN_SHEET}' no encontrada. Hojas disponibles: {available}")

    ws = wb[MAIN_SHEET]

    # Leer headers de la fila HEADER_ROW
    headers = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for col_idx, header in enumerate(row, 1):
            if header is not None:
                headers[str(header).strip()] = col_idx
        break

    # Verificar columnas requeridas
    required = [COL_CARD_ID, COL_STATUS, COL_DESCRIPCION]
    missing  = [c for c in required if c not in headers]
    if missing:
        raise ValueError(
            f"Columnas requeridas no encontradas en '{MAIN_SHEET}': {missing}\n"
            f"Columnas encontradas: {list(headers.keys())}"
        )

    # Índices de columnas (base 1 → base 0 para values_only)
    idx_card_id   = headers[COL_CARD_ID] - 1
    idx_status    = headers[COL_STATUS] - 1
    idx_desc      = headers[COL_DESCRIPCION] - 1
    idx_notas     = headers.get(COL_NOTAS, 0) - 1
    idx_val_by    = headers.get(COL_VALIDATED_BY, 0) - 1
    idx_val_dt    = headers.get(COL_VALIDATED_DT, 0) - 1

    decisions   = defaultdict(list)
    total_rows  = 0
    pending     = 0
    por_status  = defaultdict(int)

    for row in ws.iter_rows(min_row=DATA_ROW_FROM, values_only=True):
        card_id_raw = row[idx_card_id] if idx_card_id < len(row) else None
        status_raw  = row[idx_status]  if idx_status < len(row) else None

        if card_id_raw is None:
            continue

        total_rows += 1

        # Normalizar card_id
        try:
            card_id = int(str(card_id_raw).strip())
        except (ValueError, TypeError):
            continue

        status = str(status_raw).strip().lower() if status_raw else "pending_validation"
        por_status[status] += 1

        if status not in DECIDED_STATUSES:
            pending += 1
            continue

        descripcion  = str(row[idx_desc])    if idx_desc >= 0 and idx_desc < len(row) and row[idx_desc] else ""
        notas        = str(row[idx_notas])   if idx_notas > 0 and idx_notas < len(row) and row[idx_notas] else ""
        validated_by = str(row[idx_val_by])  if idx_val_by > 0 and idx_val_by < len(row) and row[idx_val_by] else ""
        validated_dt = str(row[idx_val_dt])  if idx_val_dt > 0 and idx_val_dt < len(row) and row[idx_val_dt] else ""

        decisions[card_id].append({
            "status":       status,
            "descripcion":  descripcion,
            "notas":        notas,
            "validated_by": validated_by,
            "validated_date": validated_dt,
        })

    stats = {
        "total_rows":  total_rows,
        "decididas":   total_rows - pending,
        "pending":     pending,
        "por_status":  dict(por_status),
        "cards_con_decision": len(decisions),
    }

    return dict(decisions), stats


# ============================================================
# ÍNDICE DE JSONs
# ============================================================

def build_json_index(output_dir: str) -> dict:
    """
    Escanea los JSONs de resultados y construye {card_id: json_filepath}.
    Un card_id solo puede aparecer en un lote — si hay colisión, avisa.
    """
    index = {}
    json_files = sorted(Path(output_dir).glob("lote_*_resultados.json"))

    for jf in json_files:
        with open(jf, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                print(f"   ⚠️  JSON inválido, saltando: {jf.name}")
                continue

        for card in data.get("cards", []):
            if isinstance(card, dict) and "card_id" in card:
                cid = card["card_id"]
                if cid in index:
                    print(f"   ⚠️  card_id {cid} aparece en más de un lote: {index[cid].name} y {jf.name}")
                else:
                    index[cid] = jf

    return index


# ============================================================
# DERIVAR STATUS A NIVEL CARD
# ============================================================

def derive_card_status(hallazgo_statuses: list[str]) -> tuple[str, dict]:
    """
    Dado el status de cada hallazgo de una card, devuelve:
      - status a nivel card (el más severo, o 'mixed' si hay mezcla)
      - resumen: {"confirmed_error": N, "false_positive": N, "intentional": N, "pending": N}
    """
    counts = defaultdict(int)
    for s in hallazgo_statuses:
        counts[s] += 1

    decided = {k: v for k, v in counts.items() if k in DECIDED_STATUSES}

    if not decided:
        return "pending_validation", dict(counts)

    # Si todos los decididos son el mismo tipo → ese tipo
    tipos_decididos = set(decided.keys())
    if len(tipos_decididos) == 1:
        card_status = tipos_decididos.pop()
    else:
        # Hay mezcla — usar el más severo
        max_prio   = max(STATUS_PRIORITY.get(t, 0) for t in tipos_decididos)
        card_status = next(t for t in tipos_decididos if STATUS_PRIORITY.get(t, 0) == max_prio)
        # Si hay confirmed_error pero también false_positive → 'mixed'
        if "confirmed_error" in tipos_decididos and len(tipos_decididos) > 1:
            card_status = "mixed"

    summary = {
        "confirmed_error": counts.get("confirmed_error", 0),
        "false_positive":  counts.get("false_positive",  0),
        "intentional":     counts.get("intentional",     0),
        "pending":         counts.get("pending_validation", 0),
    }
    return card_status, summary


# ============================================================
# MATCH HALLAZGO
# ============================================================

def match_hallazgo(json_hallazgo: dict, tracker_description: str,
                   fuzzy_threshold: float = 0.60) -> tuple[bool, str]:
    """
    Devuelve (matched, method) donde method indica cómo se logró el match:
      "exact"       → primeros 80 chars iguales o fragmento de 40 chars
      "fuzzy_N.NN"  → token overlap >= fuzzy_threshold
      ""            → sin match

    Estrategia en orden de prioridad:
      1. Match exacto en primeros 80 chars
      2. Fragmento: primeros 40 chars de uno dentro del otro
      3. Token overlap >= fuzzy_threshold (fallback)
    """
    if not tracker_description:
        return False, ""

    json_desc     = json_hallazgo.get("descripcion", "")
    tracker_short = tracker_description[:80].strip().lower()
    json_short    = json_desc[:80].strip().lower()

    # Match exacto en los primeros 80 chars
    if tracker_short == json_short:
        return True, "exact"

    # Match por fragmento: el tracker contiene al menos 40 chars del JSON
    if len(tracker_short) >= 40 and tracker_short[:40] in json_short:
        return True, "exact"
    if len(json_short) >= 40 and json_short[:40] in tracker_short:
        return True, "exact"

    # Fallback: token overlap
    score = _token_overlap(tracker_description, json_desc)
    if score >= fuzzy_threshold:
        return True, f"fuzzy_{score:.2f}"

    return False, ""


# ============================================================
# PROPAGAR DECISIONES A UN JSON
# ============================================================

def propagate_to_json(
    json_filepath: Path,
    card_decisions: dict,   # {card_id: [{"status", "descripcion", "notas", ...}]}
    dry_run: bool,
    diagnose: bool = False,
) -> dict:
    """
    Carga el JSON, inyecta los campos de revisión humana en las cards y hallazgos indicados.
    Devuelve un resumen de lo que se hizo.
    """
    with open(json_filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    cards_updated        = 0
    hallazgos_matched    = 0
    hallazgos_unmatched  = 0
    hallazgos_fuzzy      = 0
    modified             = False
    diagnose_items       = []

    for card in data.get("cards", []):
        if not isinstance(card, dict) or "card_id" not in card:
            continue

        cid = card["card_id"]
        if cid not in card_decisions:
            continue

        tracker_rows      = card_decisions[cid]
        hallazgo_statuses = []
        used_tracker_rows = set()
        json_hallazgos    = card.get("hallazgos", [])

        # Intentar matchear cada hallazgo del JSON con una fila del tracker
        for h_idx, h in enumerate(json_hallazgos):
            matched      = False
            match_method = ""
            for t_idx, tr in enumerate(tracker_rows):
                if t_idx in used_tracker_rows:
                    continue
                did_match, method = match_hallazgo(h, tr["descripcion"])
                if did_match:
                    match_method = method
                    # Inyectar en el hallazgo
                    h["human_review_status"]  = tr["status"]
                    h["human_review_notes"]   = tr["notas"]        if tr["notas"]        else None
                    h["validated_by"]         = tr["validated_by"] if tr["validated_by"] else None
                    h["validated_date"]       = tr["validated_date"] if tr["validated_date"] else None

                    # Registrar método de match si fue fuzzy
                    if method.startswith("fuzzy"):
                        h["match_method"] = method
                        hallazgos_fuzzy += 1

                    hallazgo_statuses.append(tr["status"])
                    used_tracker_rows.add(t_idx)
                    hallazgos_matched += 1
                    matched = True
                    modified = True
                    break

            if not matched:
                # Caso especial: card con 1 solo hallazgo y 1 sola fila tracker no usada
                unused = [t_idx for t_idx in range(len(tracker_rows)) if t_idx not in used_tracker_rows]
                if len(json_hallazgos) == 1 and len(unused) == 1:
                    t_idx = unused[0]
                    tr = tracker_rows[t_idx]
                    h["human_review_status"]  = tr["status"]
                    h["human_review_notes"]   = tr["notas"]        if tr["notas"]        else None
                    h["validated_by"]         = tr["validated_by"] if tr["validated_by"] else None
                    h["validated_date"]       = tr["validated_date"] if tr["validated_date"] else None
                    h["match_method"]         = "force_single"
                    hallazgo_statuses.append(tr["status"])
                    used_tracker_rows.add(t_idx)
                    hallazgos_matched += 1
                    hallazgos_fuzzy   += 1
                    matched = True
                    modified = True
                else:
                    h.setdefault("human_review_status", "pending_validation")
                    hallazgo_statuses.append("pending_validation")
                    hallazgos_unmatched += 1

                    if diagnose and not matched:
                        diagnose_items.append({
                            "card_id":   cid,
                            "card_name": card.get("card_name", ""),
                            "h_desc_json": h.get("descripcion", "")[:150],
                            "candidatas": [
                                {"desc": tr["descripcion"][:150], "status": tr["status"]}
                                for tr in tracker_rows
                            ],
                        })

        # Filas del tracker que no matchearon ningún hallazgo (hallazgo puede haber sido omitido por el modelo)
        unmatched_tracker = [tr for t_idx, tr in enumerate(tracker_rows) if t_idx not in used_tracker_rows]
        if unmatched_tracker:
            hallazgos_statuses_extra = [tr["status"] for tr in unmatched_tracker]
            hallazgo_statuses.extend(hallazgos_statuses_extra)

        # Derivar status y summary a nivel card
        card_status, summary = derive_card_status(hallazgo_statuses)
        card["human_review_status"]  = card_status
        card["human_review_summary"] = summary
        card["human_review_updated"] = datetime.now().strftime("%Y-%m-%d")

        cards_updated += 1
        modified = True

    if modified and not dry_run:
        with open(json_filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    return {
        "cards_updated":        cards_updated,
        "hallazgos_matched":    hallazgos_matched,
        "hallazgos_fuzzy":      hallazgos_fuzzy,
        "hallazgos_unmatched":  hallazgos_unmatched,
        "json_written":         modified and not dry_run,
        "diagnose_items":       diagnose_items,
    }


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Propaga revisiones humanas del tracker Excel a los JSONs del auditor"
    )
    parser.add_argument(
        "--inspect", action="store_true",
        help="Solo muestra estadísticas del tracker sin modificar nada"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Muestra qué se actualizaría sin escribir ningún archivo"
    )
    parser.add_argument(
        "--diagnose", action="store_true",
        help="Muestra lado a lado las descripciones de hallazgos sin match (para reparación manual)"
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  Feedback Loop — Tracker → JSONs")
    print(f"  Tracker: {TRACKER_FILE}")
    print(f"  Modo: {'INSPECT' if args.inspect else 'DRY-RUN' if args.dry_run else 'REAL'}")
    print("=" * 60)

    # 1. Cargar tracker
    print(f"\n📊 Leyendo tracker Excel...")
    try:
        decisions, stats = load_tracker(TRACKER_FILE)
    except (FileNotFoundError, ValueError) as e:
        print(f"✗ Error al leer tracker: {e}")
        return

    print(f"   Filas leídas en '{MAIN_SHEET}': {stats['total_rows']:,}")
    print(f"   Con decisión ya tomada:  {stats['decididas']:,}")
    print(f"   Pendientes de revisión:  {stats['pending']:,}")
    print(f"   Cards únicas con decisión: {stats['cards_con_decision']}")
    print()
    print("   Distribución por status:")
    for status, count in sorted(stats["por_status"].items(), key=lambda x: -x[1]):
        marker = "✅" if status in DECIDED_STATUSES else "⏳"
        print(f"     {marker} {status}: {count:,}")

    if args.inspect:
        print("\n✓ Modo --inspect: no se modificó nada.")
        return

    if stats["decididas"] == 0:
        print("\n⚠️  No hay ninguna decisión en el tracker todavía.")
        print("   Rellena la columna 'Validation Status' con:")
        print("   confirmed_error | false_positive | intentional")
        return

    # 2. Construir índice de JSONs
    print(f"\n🗂️  Indexando JSONs de resultados...")
    json_index = build_json_index(OUTPUT_DIR)
    print(f"   {len(json_index):,} cards indexadas en {len(list(Path(OUTPUT_DIR).glob('lote_*_resultados.json')))} JSONs")

    # 3. Verificar qué cards del tracker tienen JSON
    cards_con_json    = {cid for cid in decisions if cid in json_index}
    cards_sin_json    = {cid for cid in decisions if cid not in json_index}

    print(f"   Cards con decisión + JSON encontrado: {len(cards_con_json)}")
    if cards_sin_json:
        print(f"   ⚠️  Cards con decisión pero sin JSON: {len(cards_sin_json)} → {sorted(cards_sin_json)[:10]}")

    if not cards_con_json:
        print("\n✗ Ninguna card del tracker tiene JSON correspondiente. Verifica los card IDs.")
        return

    # 4. Agrupar cards por JSON file
    by_json_file = defaultdict(dict)
    for cid in cards_con_json:
        jf = json_index[cid]
        by_json_file[jf][cid] = decisions[cid]

    mode_label = "DRY-RUN — " if args.dry_run else ""
    print(f"\n{'🔍' if args.dry_run else '✍️ '} {mode_label}Propagando a {len(by_json_file)} archivos JSON...\n")

    total_cards_updated        = 0
    total_hallazgos_matched    = 0
    total_hallazgos_fuzzy      = 0
    total_hallazgos_unmatched  = 0
    total_jsons_written        = 0
    all_diagnose_items         = []

    for json_file, card_decisions in sorted(by_json_file.items()):
        result = propagate_to_json(
            json_file, card_decisions,
            dry_run=args.dry_run,
            diagnose=args.diagnose,
        )

        icon = "💾" if result["json_written"] else ("🔍" if args.dry_run else "⚠️ ")
        fuzzy_note = f" | {result['hallazgos_fuzzy']} fuzzy" if result["hallazgos_fuzzy"] else ""
        print(
            f"  {icon} {json_file.name}: "
            f"{result['cards_updated']} cards | "
            f"{result['hallazgos_matched']} matched{fuzzy_note} | "
            f"{result['hallazgos_unmatched']} sin match"
        )

        total_cards_updated       += result["cards_updated"]
        total_hallazgos_matched   += result["hallazgos_matched"]
        total_hallazgos_fuzzy     += result["hallazgos_fuzzy"]
        total_hallazgos_unmatched += result["hallazgos_unmatched"]
        all_diagnose_items.extend(result.get("diagnose_items", []))
        if result["json_written"]:
            total_jsons_written += 1

    # 5. Resumen final
    total_intentados = total_hallazgos_matched + total_hallazgos_unmatched
    match_rate = (
        total_hallazgos_matched / total_intentados * 100
        if total_intentados > 0 else 0
    )

    print(f"\n{'=' * 60}")
    if args.dry_run:
        print(f"  ✓ DRY-RUN completado — ningún archivo modificado")
    else:
        print(f"  ✓ Propagación completada")
        print(f"  💾 JSONs escritos: {total_jsons_written}")
    print(f"  🃏 Cards actualizadas: {total_cards_updated}")
    print(f"  ✅ Hallazgos matcheados: {total_hallazgos_matched}")
    if total_hallazgos_fuzzy:
        print(f"     └─ Fuzzy/force: {total_hallazgos_fuzzy} (token overlap ≥60% o force-single)")
    print(f"  ❓ Hallazgos sin match: {total_hallazgos_unmatched}")
    print(f"  📊 Match rate: {match_rate:.1f}%")
    if total_hallazgos_unmatched > 0:
        print(f"  ℹ️  Los hallazgos sin match quedan como 'pending_validation' en el JSON.")
        print(f"     Ejecuta --diagnose para ver las descripciones lado a lado.")
    print(f"{'=' * 60}")

    if args.dry_run:
        print("\n  Ejecuta sin --dry-run para aplicar los cambios.")

    # 6. Salida de diagnose
    if args.diagnose and all_diagnose_items:
        print(f"\n  {'─'*58}")
        print(f"  DIAGNOSE — {len(all_diagnose_items)} hallazgos sin match")
        print(f"  {'─'*58}")
        for item in all_diagnose_items[:30]:
            print(f"\n  Card #{item['card_id']} — {item['card_name'][:50]}")
            print(f"    JSON:   {item['h_desc_json'][:110]}")
            for c in item.get("candidatas", [])[:2]:
                print(f"    Excel:  {c['desc'][:110]}")
        if len(all_diagnose_items) > 30:
            print(f"\n  ... y {len(all_diagnose_items)-30} más (ver diagnose_pending_excel.py para reporte completo)")


if __name__ == "__main__":
    main()
