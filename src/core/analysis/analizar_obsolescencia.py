"""
Analizador de Obsolescencia de Cards — Metabase
ExampleCorp · author · 2026-03-28

¿Qué hace este script?
    1. Llama a la API de Metabase para obtener metadatos de todas las cards
       (last_used_at, created_at, creador, si está en algún dashboard)
    2. Cruza esa info con los hallazgos del auditor (view_count, patrones)
    3. Clasifica cada card en: Activa / Inactiva 6m / Inactiva 12m /
       Nunca usada / Duplicado / Con errores+inactiva
    4. Genera acciones recomendadas: Archivar / Revisar / Eliminar / Mantener
    5. Guarda el resultado en:
       - obsolescencia_cards.csv           (para revisión rápida)
       - obsolescencia_reporte.md          (reporte legible)
       - Agrega hoja al tracker Excel      (tracker_auditoria_metabase.xlsx)

Prerequisito:
    METABASE_SESSION=<tu_session_token> en config/api_key.env
    (DevTools → Application → Cookies → metabase.SESSION)

Uso:
    python analizar_obsolescencia.py --dry-run      # solo muestra estadísticas, no escribe
    python analizar_obsolescencia.py                # corre completo
"""

import os
import sys
import json
import csv
import time
import argparse
import requests
from datetime import datetime, timezone, timedelta
from pathlib import Path

# ──────────────────────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────────────────────
_ROOT         = Path(os.path.dirname(__file__)).resolve().parent.parent.parent
sys.path.insert(0, str(_ROOT))
from src.utils.metabase_client import METABASE_URL, get_session_token
RESULTS_DIR   = str(_ROOT / "data" / "processed" / "resultados")
TRACKER_PATH  = str(_ROOT / "data" / "raw" / "tracker_auditoria_metabase.xlsx")
OUT_CSV       = str(_ROOT / "data" / "raw" / "obsolescencia_cards.csv")
OUT_MD        = str(_ROOT / "docs" / "obsolescencia_reporte.md")

NOW           = datetime.now(timezone.utc)
THRESHOLD_6M  = NOW - timedelta(days=182)
THRESHOLD_12M = NOW - timedelta(days=365)
THRESHOLD_18M = NOW - timedelta(days=548)

# Patrones que nos dicen que la card tiene un error confirmado o fix listo
P9A_IDS = {198,114,197,238,596,97,139,199,451,520,377,399,437,634,665,527,
           394,395,462,493,495,501,200,419,189,872,874,878,880,912,973}
P9B_IDS = {277,828,829,830,831,1054,820,2892}
P3_IDS  = {50,52,1136,1098,1375}

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

load_session_token = get_session_token

def get_headers(token):
    return {"X-Metabase-Session": token, "Content-Type": "application/json"}

def parse_date(s):
    """Parsea fecha ISO de Metabase → datetime con tz."""
    if not s:
        return None
    s = s.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except Exception:
        try:
            return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
        except Exception:
            return None

def days_ago(dt):
    if not dt:
        return None
    delta = NOW - dt
    return delta.days

def load_findings_from_json():
    """Carga los hallazgos del auditor: {card_id: {views, patterns, severidades}}"""
    card_data = {}
    if not os.path.exists(RESULTS_DIR):
        print(f"⚠️  No se encontró carpeta de resultados: {RESULTS_DIR}")
        return card_data

    for fname in sorted(os.listdir(RESULTS_DIR)):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(RESULTS_DIR, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        cards = data if isinstance(data, list) else data.get("cards", data.get("results", []))
        for card in cards:
            cid = str(card.get("card_id") or card.get("id") or "")
            if not cid:
                continue
            views    = int(card.get("views") or card.get("vistas") or 0)
            findings = card.get("findings") or card.get("hallazgos") or []
            patterns = [str(f.get("pattern") or f.get("patron") or "P99").upper()
                        for f in findings]
            sevs     = [str(f.get("severity") or f.get("severidad") or "baja").lower()
                        for f in findings]
            if cid not in card_data:
                card_data[cid] = {"views": views, "patterns": [], "severidades": []}
            card_data[cid]["views"] = max(card_data[cid]["views"], views)
            card_data[cid]["patterns"].extend(patterns)
            card_data[cid]["severidades"].extend(sevs)

    return card_data

def classify_card(card, audit_data, now=NOW):
    """Retorna (categoria, accion, razon) para una card."""
    cid       = str(card.get("id", ""))
    name      = card.get("name") or ""
    views     = card.get("view_count") or 0
    last_used = parse_date(card.get("last_used_at"))
    created   = parse_date(card.get("created_at"))
    in_dash   = bool(card.get("dashboard_count", 0))
    archived  = card.get("archived", False)

    name_lower = name.lower()
    is_duplicate = any(x in name_lower for x in [
        "duplicate", "- dup", "copy", " v2", " v3", "deprecated", "old ",
        "(1)", "(2)", "backup", "test", "prueba",
    ])
    # Verificar también si el ID terminó en patrones de duplicado conocidos
    audit = audit_data.get(cid, {})
    has_alta = "alta" in audit.get("severidades", [])
    has_fix_ready = (int(cid) in P9A_IDS or int(cid) in P9B_IDS or int(cid) in P3_IDS) if cid.isdigit() else False
    has_findings  = bool(audit.get("patterns"))

    # Categorías de uso
    if last_used is None and views == 0:
        uso = "NUNCA_USADA"
        dias = None
    elif last_used is None and views > 0:
        uso = "SIN_FECHA_PERO_CON_VISTAS"
        dias = None
    else:
        dias = days_ago(last_used)
        if dias is not None:
            if dias > 548:
                uso = "INACTIVA_18M+"
            elif dias > 365:
                uso = "INACTIVA_12M"
            elif dias > 182:
                uso = "INACTIVA_6M"
            else:
                uso = "ACTIVA"
        else:
            uso = "SIN_FECHA"

    # Definir acción
    if archived:
        return "ARCHIVADA", "Ya archivada", "Ya fue archivada — solo registrar", dias

    if uso == "NUNCA_USADA":
        if in_dash:
            accion = "REVISAR_DASHBOARD"
            razon  = "Nunca usada pero está en un dashboard — verificar si el dashboard se usa"
        elif is_duplicate:
            accion = "ELIMINAR"
            razon  = "Nunca usada + nombre indica duplicado/copia"
        elif has_alta and has_findings:
            accion = "ELIMINAR"
            razon  = "Nunca usada + tiene errores de severidad alta — no vale la pena corregir"
        else:
            accion = "ARCHIVAR"
            razon  = "Nunca usada — archivar para limpiar el inventario"

    elif uso in ("INACTIVA_18M+", "INACTIVA_12M"):
        if in_dash:
            accion = "REVISAR_DASHBOARD"
            razon  = f"Inactiva {dias} días pero en dashboard — validar si dashboard sigue activo"
        elif is_duplicate:
            accion = "ELIMINAR"
            razon  = f"Inactiva {dias} días + nombre indica duplicado"
        elif has_fix_ready:
            accion = "ARCHIVAR_TRAS_FIX"
            razon  = f"Inactiva {dias} días + tiene fix pendiente — aplicar fix o archivar"
        else:
            accion = "ARCHIVAR"
            razon  = f"Sin uso en {dias} días"

    elif uso == "INACTIVA_6M":
        if is_duplicate:
            accion = "REVISAR_ELIMINAR"
            razon  = "Inactiva 6m + nombre indica duplicado"
        elif has_alta:
            accion = "REVISAR_CON_DUENO"
            razon  = f"Inactiva {dias} días + tiene errores altos — preguntar si sigue siendo necesaria"
        else:
            accion = "REVISAR_CON_DUENO"
            razon  = f"Inactiva {dias} días — preguntar al dueño si sigue siendo relevante"

    elif uso == "SIN_FECHA_PERO_CON_VISTAS":
        if is_duplicate:
            accion = "REVISAR_ELIMINAR"
            razon  = "Tiene vistas pero nombre indica duplicado — revisar con dueño"
        elif has_fix_ready:
            accion = "APLICAR_FIX"
            razon  = "Tiene vistas + fix verificado listo — priorizar corrección"
        else:
            accion = "MANTENER"
            razon  = "Tiene vistas — mantener"

    elif uso == "ACTIVA":
        if is_duplicate:
            accion = "REVISAR_ELIMINAR"
            razon  = "Activa pero nombre dice duplicado — confirmar cuál es la versión oficial"
        elif has_fix_ready:
            accion = "APLICAR_FIX"
            razon  = "Activa + fix verificado listo — ALTA PRIORIDAD corregir"
        else:
            accion = "MANTENER"
            razon  = "Card activa y sin problemas críticos"

    else:
        accion = "REVISAR_CON_DUENO"
        razon  = "Datos insuficientes para clasificar"

    return uso, accion, razon, dias

# ──────────────────────────────────────────────────────────────────────────────
# Fetch de Metabase
# ──────────────────────────────────────────────────────────────────────────────

def fetch_all_cards(token):
    """Obtiene todas las cards de Metabase con sus metadatos."""
    headers = get_headers(token)
    url     = f"{METABASE_URL}/api/card"
    print(f"📡 Llamando a {url} ...")
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        print(f"   → {len(data)} cards recibidas")
        return data
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            print("❌ Error 401 — sesión expirada. Obtén un nuevo METABASE_SESSION.")
        elif e.response.status_code == 403:
            print("❌ Error 403 — sin permisos para /api/card.")
        else:
            print(f"❌ Error HTTP {e.response.status_code}: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error de conexión: {e}")
        sys.exit(1)

def fetch_dashboard_cards(token):
    """
    Obtiene qué cards están en dashboards.
    Retorna: {card_id: [dashboard_name, ...]}
    """
    headers = get_headers(token)
    url     = f"{METABASE_URL}/api/dashboard"
    cards_in_dash = {}
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        dashboards = resp.json()
        print(f"   → {len(dashboards)} dashboards encontrados")

        for i, dash in enumerate(dashboards):
            did   = dash.get("id")
            dname = dash.get("name", f"Dashboard {did}")
            # Obtener detalle del dashboard para ver sus cards
            try:
                dr = requests.get(f"{METABASE_URL}/api/dashboard/{did}",
                                  headers=headers, timeout=30)
                dr.raise_for_status()
                ddata = dr.json()
                for dc in ddata.get("dashcards", ddata.get("ordered_cards", [])):
                    cid = str(dc.get("card_id") or "")
                    if cid and cid != "None":
                        cards_in_dash.setdefault(cid, []).append(dname)
            except Exception:
                pass
            # Rate limiting suave
            if i % 20 == 0 and i > 0:
                time.sleep(0.5)
                print(f"   ... {i}/{len(dashboards)} dashboards procesados")

        print(f"   → {len(cards_in_dash)} cards tienen al menos un dashboard")
        return cards_in_dash
    except Exception as e:
        print(f"⚠️  No se pudo obtener dashboards: {e}. Continuando sin esa info.")
        return {}

# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main(dry_run=False):
    print("\n" + "="*60)
    print("  ANALIZADOR DE OBSOLESCENCIA — Cards Metabase ExampleCorp")
    print(f"  Fecha de corte: {NOW.strftime('%Y-%m-%d')}")
    print("="*60 + "\n")

    # 1. Cargar token
    token = load_session_token()
    print(f"✅ Token cargado desde {ENV_FILE}")

    # 2. Cargar hallazgos del auditor
    print("\n📂 Cargando hallazgos del auditor GPT-4o...")
    audit_data = load_findings_from_json()
    print(f"   → {len(audit_data)} cards con hallazgos en JSONs")

    # 3. Obtener cards de Metabase
    print("\n📡 Obteniendo metadatos de Metabase...")
    mb_cards = fetch_all_cards(token)

    # 4. Obtener info de dashboards
    print("\n📊 Obteniendo info de dashboards...")
    cards_in_dash = fetch_dashboard_cards(token)

    # 5. Enriquecer cards con info de dashboards
    for card in mb_cards:
        cid = str(card.get("id", ""))
        card["dashboard_count"] = len(cards_in_dash.get(cid, []))
        card["dashboards"]      = ", ".join(cards_in_dash.get(cid, []))

    # 6. Clasificar todas las cards
    print("\n🔍 Clasificando cards...")
    results = []
    for card in mb_cards:
        cid   = str(card.get("id", ""))
        audit = audit_data.get(cid, {})
        uso, accion, razon, dias = classify_card(card, audit_data)

        last_used_raw = card.get("last_used_at") or ""
        created_raw   = card.get("created_at") or ""
        creator       = (card.get("creator") or {})
        creator_name  = creator.get("common_name") or creator.get("email") or "Desconocido"

        # Datos de hallazgos del auditor
        patterns_str  = ", ".join(sorted(set(audit.get("patterns", [])))) or "—"
        has_alta      = "alta" in audit.get("severidades", [])
        fix_ready     = cid.isdigit() and (
            int(cid) in P9A_IDS or int(cid) in P9B_IDS or int(cid) in P3_IDS
        )

        results.append({
            "card_id":         cid,
            "nombre":          card.get("name", ""),
            "views":           card.get("view_count", 0),
            "uso_categoria":   uso,
            "dias_inactiva":   dias if dias is not None else "",
            "last_used_at":    last_used_raw[:10] if last_used_raw else "Nunca",
            "created_at":      created_raw[:10] if created_raw else "",
            "creator":         creator_name,
            "en_dashboards":   card["dashboard_count"],
            "dashboards":      card["dashboards"],
            "accion":          accion,
            "razon":           razon,
            "patrones_audit":  patterns_str,
            "tiene_error_alta": "Sí" if has_alta else "No",
            "fix_listo":       "Sí" if fix_ready else "No",
            "collection":      (card.get("collection") or {}).get("name") or "Sin colección",
            "archived":        "Sí" if card.get("archived") else "No",
        })

    # 7. Estadísticas
    total          = len(results)
    nunca          = sum(1 for r in results if r["uso_categoria"] == "NUNCA_USADA")
    inactiva_18    = sum(1 for r in results if r["uso_categoria"] == "INACTIVA_18M+")
    inactiva_12    = sum(1 for r in results if r["uso_categoria"] == "INACTIVA_12M")
    inactiva_6     = sum(1 for r in results if r["uso_categoria"] == "INACTIVA_6M")
    activas        = sum(1 for r in results if r["uso_categoria"] == "ACTIVA")
    archivar       = sum(1 for r in results if r["accion"] in ("ARCHIVAR", "ARCHIVAR_TRAS_FIX"))
    eliminar       = sum(1 for r in results if r["accion"] in ("ELIMINAR", "REVISAR_ELIMINAR"))
    revisar        = sum(1 for r in results if "REVISAR" in r["accion"])
    aplicar_fix    = sum(1 for r in results if r["accion"] == "APLICAR_FIX")
    duplicados     = sum(1 for r in results if "DUPLICADO" in r.get("razon","").upper() or
                         any(x in (r["nombre"] or "").lower()
                             for x in ["duplicate","- dup","copy","backup"]))
    en_dash_count  = sum(1 for r in results if int(r["en_dashboards"]) > 0)

    print(f"\n{'='*60}")
    print(f"  RESULTADOS — {total} cards analizadas")
    print(f"{'='*60}")
    print(f"  📊 CATEGORÍAS DE USO:")
    print(f"     ✅ Activas (usadas en <6 meses)    : {activas:4d} ({100*activas//total}%)")
    print(f"     🟡 Inactivas 6-12 meses            : {inactiva_6:4d} ({100*inactiva_6//total}%)")
    print(f"     🟠 Inactivas 12-18 meses           : {inactiva_12:4d} ({100*inactiva_12//total}%)")
    print(f"     🔴 Inactivas >18 meses             : {inactiva_18:4d} ({100*inactiva_18//total}%)")
    print(f"     ⚫ Nunca usadas (0 vistas, sin fecha): {nunca:4d} ({100*nunca//total}%)")
    print(f"")
    print(f"  🎯 ACCIONES RECOMENDADAS:")
    print(f"     🗑️  Candidatas a ARCHIVAR           : {archivar}")
    print(f"     ❌ Candidatas a ELIMINAR            : {eliminar}")
    print(f"     🔁 Candidatas a APLICAR FIX         : {aplicar_fix}")
    print(f"     👥 Requieren revisión con dueño     : {revisar}")
    print(f"")
    print(f"  📌 DATOS EXTRA:")
    print(f"     Cards en dashboards                : {en_dash_count}")
    print(f"     Cards con nombre 'Duplicate/Copy'  : {duplicados}")
    print(f"{'='*60}\n")

    if dry_run:
        print("🔵 Modo --dry-run. No se escriben archivos.")
        return results

    # 8. Guardar CSV
    if results:
        with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
            writer.writeheader()
            writer.writerows(results)
        print(f"✅ CSV guardado: {OUT_CSV}")

    # 9. Guardar reporte Markdown
    _save_markdown(results, total, nunca, inactiva_18, inactiva_12, inactiva_6,
                   activas, archivar, eliminar, aplicar_fix, duplicados, en_dash_count)

    # 10. Agregar hoja al tracker Excel
    _add_excel_sheet(results)

    print(f"\n🎉 ¡Listo! Archivos generados:")
    print(f"   → {OUT_CSV}")
    print(f"   → {OUT_MD}")
    print(f"   → {TRACKER_PATH}  (nueva hoja '🗑️ Obsolescencia')")

    return results


def _save_markdown(results, total, nunca, inactiva_18, inactiva_12, inactiva_6,
                   activas, archivar, eliminar, aplicar_fix, duplicados, en_dash_count):
    ahora = NOW.strftime("%d/%m/%Y")
    lines = [
        f"# Reporte de Obsolescencia — Cards Metabase",
        f"**ExampleCorp · author · {ahora}**",
        f"",
        f"> Este reporte cruza los datos de uso de Metabase (last_used_at) con los hallazgos del auditor GPT-4o.",
        f"> Objetivo: identificar cards que ya no son necesarias y limpiar el inventario de 3,014 cards.",
        f"",
        f"---",
        f"",
        f"## Resumen de Uso",
        f"",
        f"| Categoría | Cards | % del total |",
        f"|-----------|-------|------------|",
        f"| ✅ Activas (usadas en <6 meses) | {activas} | {100*activas//total}% |",
        f"| 🟡 Inactivas 6-12 meses | {inactiva_6} | {100*inactiva_6//total}% |",
        f"| 🟠 Inactivas 12-18 meses | {inactiva_12} | {100*inactiva_12//total}% |",
        f"| 🔴 Inactivas >18 meses | {inactiva_18} | {100*inactiva_18//total}% |",
        f"| ⚫ Nunca usadas | {nunca} | {100*nunca//total}% |",
        f"| **TOTAL** | **{total}** | **100%** |",
        f"",
        f"---",
        f"",
        f"## Acciones Recomendadas",
        f"",
        f"| Acción | Cards |",
        f"|--------|-------|",
        f"| 🗑️ Archivar (inactivas, sin impacto) | {archivar} |",
        f"| ❌ Eliminar (duplicados + nunca usadas) | {eliminar} |",
        f"| 🔁 Aplicar fix y mantener | {aplicar_fix} |",
        f"| 📋 Cards en dashboards activos (NO tocar sin revisar) | {en_dash_count} |",
        f"",
        f"---",
        f"",
        f"## Top Candidatas a Eliminar (duplicados + nunca usadas + con errores)",
        f"",
        f"| Card ID | Nombre | Días inactiva | Acción | Razón |",
        f"|---------|--------|--------------|--------|-------|",
    ]

    top_eliminar = [r for r in results if r["accion"] in ("ELIMINAR", "REVISAR_ELIMINAR")]
    top_eliminar.sort(key=lambda x: int(x["dias_inactiva"] or 9999), reverse=True)
    for r in top_eliminar[:30]:
        lines.append(f"| {r['card_id']} | {r['nombre'][:50]} | {r['dias_inactiva']} | {r['accion']} | {r['razon'][:60]} |")

    lines += [
        f"",
        f"---",
        f"",
        f"## Top Candidatas a Archivar (inactivas >12 meses)",
        f"",
        f"| Card ID | Nombre | Último uso | Días | Creator | En Dashboard |",
        f"|---------|--------|-----------|------|---------|-------------|",
    ]

    top_archivar = [r for r in results if r["accion"] == "ARCHIVAR"
                    and r["uso_categoria"] in ("INACTIVA_12M","INACTIVA_18M+")]
    top_archivar.sort(key=lambda x: int(x["dias_inactiva"] or 0), reverse=True)
    for r in top_archivar[:30]:
        lines.append(f"| {r['card_id']} | {r['nombre'][:45]} | {r['last_used_at']} | {r['dias_inactiva']} | {r['creator'][:20]} | {'✅' if int(r['en_dashboards'])>0 else '—'} |")

    lines += [
        f"",
        f"---",
        f"",
        f"## Cards Activas con Fix Listo (máxima prioridad)",
        f"",
        f"| Card ID | Nombre | Vistas | Fix | Último uso |",
        f"|---------|--------|--------|-----|-----------|",
    ]

    activas_fix = [r for r in results if r["accion"] == "APLICAR_FIX" and r["fix_listo"] == "Sí"]
    activas_fix.sort(key=lambda x: int(x["views"] or 0), reverse=True)
    for r in activas_fix[:20]:
        lines.append(f"| {r['card_id']} | {r['nombre'][:45]} | {r['views']:,} | {r['patrones_audit']} | {r['last_used_at']} |")

    lines += [
        f"",
        f"---",
        f"",
        f"## Proceso Recomendado de Limpieza",
        f"",
        f"### Fase 1 — Fix de errores en cards activas (aprobación data_owner pendiente)",
        f"```bash",
        f"python fix_masivo.py --dry-run",
        f"python fix_masivo.py --solo-p9a   # 31 cards array viejo → 9 status",
        f"python fix_masivo.py --solo-p9b   # 8 cards falta Chargeback_unir",
        f"python fix_masivo.py --solo-p3    # 5 cards renombrar título",
        f"```",
        f"",
        f"### Fase 2 — Archivar cards inactivas >12 meses (sin impacto en dashboards)",
        f"Usar la API de Metabase o la UI de Metabase:",
        f"- En Metabase UI: abrir card → menú '...' → 'Archive'",
        f"- Vía API: `PUT /api/card/{{id}}` con `{{\"archived\": true}}`",
        f"",
        f"### Fase 3 — Eliminar duplicados verificados",
        f"Solo eliminar cuando se confirme que el original existe y está correcto.",
        f"- Búsqueda de duplicados: buscar en Metabase el nombre sin '- Duplicate'",
        f"- Verificar que el original tiene el mismo SQL corregido",
        f"- Eliminar la copia con: `DELETE /api/card/{{id}}` (irreversible)",
        f"",
        f"---",
        f"*Generado por analizar_obsolescencia.py · {ahora}*",
    ]

    with open(OUT_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"✅ Reporte MD guardado: {OUT_MD}")


def _add_excel_sheet(results):
    """Agrega (o reemplaza) la hoja '🗑️ Obsolescencia' en el tracker Excel."""
    if not os.path.exists(TRACKER_PATH):
        print(f"⚠️  No se encontró el tracker: {TRACKER_PATH}")
        print("   Generando solo CSV y MD.")
        return

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  openpyxl no instalado. Solo se generó CSV y MD.")
        return

    wb = load_workbook(TRACKER_PATH)

    # Eliminar hoja si ya existe
    if "🗑️ Obsolescencia" in wb.sheetnames:
        del wb["🗑️ Obsolescencia"]

    ws = wb.create_sheet("🗑️ Obsolescencia")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    THIN   = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")
    WRAP   = Alignment(wrap_text=True, vertical="top")

    # Colores por acción
    ACCION_COLORS = {
        "ARCHIVAR":          "FFF2CC",
        "ARCHIVAR_TRAS_FIX": "FFF2CC",
        "ELIMINAR":          "FFD7D7",
        "REVISAR_ELIMINAR":  "FFDDB3",
        "APLICAR_FIX":       "C6EFCE",
        "REVISAR_CON_DUENO": "EBF3FB",
        "REVISAR_DASHBOARD": "E2EFDA",
        "MANTENER":          "FFFFFF",
        "YA_ARCHIVADA":      "D9D9D9",
        "Ya archivada":      "D9D9D9",
    }
    USO_COLORS = {
        "NUNCA_USADA":            "FF4D4D",
        "INACTIVA_18M+":          "FF8C00",
        "INACTIVA_12M":           "FFA500",
        "INACTIVA_6M":            "FFD700",
        "ACTIVA":                 "70AD47",
        "SIN_FECHA_PERO_CON_VISTAS": "A9D18E",
    }

    cols = [
        ("Card ID", 10), ("Nombre", 42), ("Vistas", 9),
        ("Categoría Uso", 20), ("Días Inactiva", 14),
        ("Último Uso", 13), ("Creada", 11), ("Creator", 22),
        ("En Dashboards", 14), ("Acción", 22), ("Razón", 48),
        ("Patrones Auditoría", 26), ("Error Alta", 11), ("Fix Listo", 10),
        ("Colección", 24),
    ]

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    tc = ws["A1"]
    tc.value = f"🗑️ ANÁLISIS DE OBSOLESCENCIA — {len(results):,} cards · Corte {NOW.strftime('%d/%m/%Y')}"
    tc.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill  = PatternFill("solid", fgColor="1F4E79")
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Cabeceras
    for ci, (cname, cw) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=cname)
        c.font   = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill   = PatternFill("solid", fgColor="2E75B6")
        c.border = BORDER
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = cw

    # Datos
    sorted_results = sorted(results,
                             key=lambda x: (
                                 ["ELIMINAR","ARCHIVAR","REVISAR_ELIMINAR","ARCHIVAR_TRAS_FIX",
                                  "APLICAR_FIX","REVISAR_CON_DUENO","REVISAR_DASHBOARD",
                                  "MANTENER","Ya archivada","ARCHIVADA"].index(x["accion"])
                                 if x["accion"] in ["ELIMINAR","ARCHIVAR","REVISAR_ELIMINAR",
                                 "ARCHIVAR_TRAS_FIX","APLICAR_FIX","REVISAR_CON_DUENO",
                                 "REVISAR_DASHBOARD","MANTENER","Ya archivada","ARCHIVADA"]
                                 else 9,
                                 -(int(x["dias_inactiva"]) if str(x["dias_inactiva"]).isdigit() else 0)
                             ))

    for ri, r in enumerate(sorted_results, 3):
        bg_accion = ACCION_COLORS.get(r["accion"], "FFFFFF")
        uso_color = USO_COLORS.get(r["uso_categoria"], "FFFFFF")

        vals = [
            r["card_id"], r["nombre"], r["views"] or 0,
            r["uso_categoria"], r["dias_inactiva"] or "",
            r["last_used_at"], r["created_at"], r["creator"],
            r["en_dashboards"] or 0, r["accion"], r["razon"],
            r["patrones_audit"], r["tiene_error_alta"], r["fix_listo"],
            r["collection"],
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Arial", size=9)
            c.border = BORDER
            # Color por columna
            if ci == 4:   # uso
                c.fill = PatternFill("solid", fgColor=uso_color + "88")
            elif ci == 10: # accion
                c.fill = PatternFill("solid", fgColor=bg_accion)
                c.font = Font(name="Arial", size=9, bold=True)
            else:
                c.fill = PatternFill("solid", fgColor="F7FBFF" if ri % 2 == 0 else "FFFFFF")
            c.alignment = CENTER if ci in (1, 3, 4, 5, 6, 7, 9, 10, 13, 14) else WRAP
        ws.row_dimensions[ri].height = 30

    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{ri}"

    # Reordenar: poner esta hoja antes de "🔍 Todos los hallazgos"
    sheet_names = wb.sheetnames
    if "🗑️ Obsolescencia" in sheet_names and "🔍 Todos los hallazgos" in sheet_names:
        target_idx = sheet_names.index("🔍 Todos los hallazgos")
        current_idx = sheet_names.index("🗑️ Obsolescencia")
        wb.move_sheet("🗑️ Obsolescencia", offset=target_idx - current_idx)

    wb.save(TRACKER_PATH)
    print(f"✅ Hoja 'Obsolescencia' agregada al tracker: {TRACKER_PATH}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analiza obsolescencia de cards Metabase")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo muestra estadísticas, no escribe archivos")
    args = parser.parse_args()
    main(dry_run=args.dry_run)
