"""
diagnose_pending_excel.py
Diagnóstica el estado de los ~365 hallazgos pending_validation en el tracker Excel:

  Grupo A — 107 "sin match": filas Excel con decisión tomada que propagate_reviews.py
            no pudo enlazar con ningún hallazgo JSON. Muestra ambas descripciones
            lado a lado para entender la causa del mismatch.

  Grupo B — ~258 "pending con match": filas Excel aún sin decisión. Para cada una
            muestra el estado actual en JSON (puede que auto_classify ya lo decidió)
            → identifica cuáles pueden sincronizarse del JSON al Excel sin revisión manual.

  Grupo C — filas Excel sin card_id en ningún JSON (cards eliminadas de Metabase).

Uso:
    python diagnose_pending_excel.py                # reporte JSON + resumen consola
    python diagnose_pending_excel.py --verbose      # muestra cada item en consola
    python diagnose_pending_excel.py --grupo a      # solo grupo A (sin match)
    python diagnose_pending_excel.py --grupo b      # solo grupo B (pending con match)
    python diagnose_pending_excel.py --stats        # solo estadísticas, sin guardar

Salida: data/processed/resultados/pending_excel_report.json
"""

import sys
import json
import argparse
import re
import openpyxl
from pathlib import Path
from collections import defaultdict
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ============================================================
# CONFIG
# ============================================================

_ROOT        = Path(__file__).resolve().parent.parent.parent.parent
TRACKER_FILE = _ROOT / "data" / "raw" / "tracker_auditoria_metabase.xlsx"
RESULTS_DIR  = _ROOT / "data" / "processed" / "resultados"
REPORT_PATH  = RESULTS_DIR / "pending_excel_report.json"

MAIN_SHEET    = "🔍 Todos los hallazgos"
HEADER_ROW    = 2
DATA_ROW_FROM = 3

COL_CARD_ID  = "Card ID"
COL_STATUS   = "Validation Status"
COL_DESC     = "Descripción Hallazgo"
COL_PATRON   = "Patrón"
COL_NOTAS    = "Notas Verificación"

DECIDED_STATUSES = {"confirmed_error", "false_positive", "intentional"}
PENDING          = "pending_validation"


# ============================================================
# CARGA DE DATOS
# ============================================================

def load_excel_rows() -> tuple[list[dict], list[dict]]:
    """
    Lee el tracker Excel y devuelve:
      - decided_rows: filas con decisión (confirmed_error / false_positive / intentional)
      - pending_rows: filas con pending_validation o sin status
    """
    wb = openpyxl.load_workbook(str(TRACKER_FILE), read_only=True, data_only=True)
    ws = wb[MAIN_SHEET]

    # Leer headers
    headers = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for idx, h in enumerate(row, 1):
            if h is not None:
                headers[str(h).strip()] = idx - 1  # base 0
        break

    def get(row, col):
        idx = headers.get(col, -1)
        if idx < 0 or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    decided_rows = []
    pending_rows = []

    for row in ws.iter_rows(min_row=DATA_ROW_FROM, values_only=True):
        card_id_raw = get(row, COL_CARD_ID)
        if not card_id_raw:
            continue
        try:
            card_id = int(card_id_raw)
        except ValueError:
            continue

        status_raw = get(row, COL_STATUS) or ""
        status = status_raw.lower()
        desc   = get(row, COL_DESC) or ""
        patron = get(row, COL_PATRON) or ""
        notas  = get(row, COL_NOTAS) or ""

        entry = {
            "card_id": card_id,
            "status":  status,
            "desc":    desc,
            "patron":  patron,
            "notas":   notas,
        }

        if status in DECIDED_STATUSES:
            decided_rows.append(entry)
        else:
            pending_rows.append(entry)

    return decided_rows, pending_rows


def load_json_index() -> dict:
    """Construye {card_id: {"hallazgos": [...], "card_name": ..., "vistas": ...}}."""
    index = {}
    for jf in sorted(RESULTS_DIR.glob("lote_*_resultados.json")):
        try:
            with open(jf, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue

        for card in data.get("cards", []):
            cid = card.get("card_id")
            if cid is None:
                continue
            if cid not in index:
                index[cid] = {
                    "card_name": card.get("card_name", ""),
                    "vistas":    card.get("vistas", 0) or 0,
                    "coleccion": card.get("coleccion", "") or "",
                    "hallazgos": card.get("hallazgos", []),
                }
    return index


# ============================================================
# MATCHING
# ============================================================

def token_overlap(a: str, b: str) -> float:
    """Jaccard sobre tokens de palabras."""
    ta = set(re.findall(r"\w+", a.lower()))
    tb = set(re.findall(r"\w+", b.lower()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def match_exact(desc_excel: str, desc_json: str) -> bool:
    """Match exacto en primeros 80 chars (lógica original)."""
    a = desc_excel[:80].strip().lower()
    b = desc_json[:80].strip().lower()
    if a == b:
        return True
    if len(a) >= 40 and a[:40] in b:
        return True
    if len(b) >= 40 and b[:40] in a:
        return True
    return False


def match_fuzzy(desc_excel: str, desc_json: str, threshold: float = 0.60) -> bool:
    return token_overlap(desc_excel, desc_json) >= threshold


def find_best_match(desc_excel: str, json_hallazgos: list[dict]) -> tuple[int, str, float]:
    """
    Busca el hallazgo JSON más parecido a desc_excel.
    Devuelve (índice, método_match, score) o (-1, "none", 0) si no hay match.
    """
    best_idx   = -1
    best_score = 0.0
    best_method = "none"

    for i, h in enumerate(json_hallazgos):
        d = h.get("descripcion", "")
        if match_exact(desc_excel, d):
            return i, "exact", 1.0
        score = token_overlap(desc_excel, d)
        if score > best_score:
            best_score = score
            best_idx   = i
            best_method = f"fuzzy_{score:.2f}"

    if best_score >= 0.60:
        return best_idx, best_method, best_score
    return -1, "none", 0.0


# ============================================================
# ANÁLISIS
# ============================================================

def analyze(verbose: bool = False, grupo: str = "all") -> dict:
    print("  Cargando tracker Excel...")
    decided_rows, pending_rows = load_excel_rows()

    print("  Cargando JSONs de resultados...")
    json_index = load_json_index()

    # ─────────────────────────────────────────────
    # GRUPO A: filas decididas en Excel sin match JSON
    # (propagate_reviews.py no pudo enlazarlas)
    # ─────────────────────────────────────────────
    grupo_a = []

    # Agrupar filas decididas por card_id
    decided_by_card = defaultdict(list)
    for r in decided_rows:
        decided_by_card[r["card_id"]].append(r)

    for card_id, rows in sorted(decided_by_card.items()):
        card_info = json_index.get(card_id)

        if card_info is None:
            # Card no existe en JSONs
            for r in rows:
                grupo_a.append({
                    "card_id":   card_id,
                    "card_name": "?? no encontrada en JSONs",
                    "vistas":    0,
                    "causa":     "card_not_in_json",
                    "desc_excel": r["desc"][:200],
                    "desc_json_candidatas": [],
                    "patron":    r["patron"],
                    "status_excel": r["status"],
                    "mejor_match_score": 0.0,
                    "mejor_match_metodo": "none",
                })
            continue

        json_hallazgos = card_info["hallazgos"]

        for r in rows:
            # ¿Ya matcheó con algún hallazgo?
            idx, method, score = find_best_match(r["desc"], json_hallazgos)

            if idx >= 0 and method == "exact":
                # Matcheó — no es "sin match"
                continue

            # No matcheó con lógica original — es candidato a Grupo A
            # (si idx >= 0 por fuzzy, podría repararse)
            candidatas = [
                {
                    "idx":     i,
                    "desc":    h.get("descripcion", "")[:200],
                    "status_json": h.get("human_review_status", "pending_validation"),
                    "overlap": round(token_overlap(r["desc"], h.get("descripcion", "")), 3),
                }
                for i, h in enumerate(json_hallazgos)
            ]
            candidatas.sort(key=lambda x: -x["overlap"])

            causa = "sin_match"
            if len(json_hallazgos) == 1 and len(rows) == 1:
                causa = "single_hallazgo_force_match_possible"
            elif idx >= 0:
                causa = f"fuzzy_reparable_{score:.2f}"

            entry = {
                "card_id":   card_id,
                "card_name": card_info["card_name"],
                "vistas":    card_info["vistas"],
                "causa":     causa,
                "desc_excel": r["desc"][:200],
                "desc_json_candidatas": candidatas[:3],
                "patron":    r["patron"],
                "status_excel": r["status"],
                "mejor_match_score":  round(score, 3),
                "mejor_match_metodo": method,
            }
            grupo_a.append(entry)

            if verbose and (grupo in ("all", "a")):
                print(f"\n  [GRUPO A] card #{card_id} ({card_info['vistas']:,}v) — {card_info['card_name'][:50]}")
                print(f"    Causa:  {causa}")
                print(f"    Excel:  {r['desc'][:100]}...")
                if candidatas:
                    print(f"    JSON:   {candidatas[0]['desc'][:100]}...")
                    print(f"    Overlap:{candidatas[0]['overlap']:.2f}")

    # ─────────────────────────────────────────────
    # GRUPO B: filas pending en Excel
    # ─────────────────────────────────────────────
    grupo_b_sin_json  = []  # card no existe en JSONs
    grupo_b_json_decidido = []  # JSON ya tiene decisión → se puede auto-sync
    grupo_b_json_pending  = []  # JSON también pendiente → necesita SQL/data_owner

    pending_by_card = defaultdict(list)
    for r in pending_rows:
        pending_by_card[r["card_id"]].append(r)

    for card_id, rows in sorted(pending_by_card.items()):
        card_info = json_index.get(card_id)

        if card_info is None:
            for r in rows:
                grupo_b_sin_json.append({
                    "card_id": card_id,
                    "patron":  r["patron"],
                    "desc":    r["desc"][:200],
                })
            continue

        json_hallazgos = card_info["hallazgos"]

        for r in rows:
            # Buscar hallazgo matching en JSON
            idx, method, score = find_best_match(r["desc"], json_hallazgos)

            json_status = "no_match"
            json_desc   = ""
            json_notes  = ""

            if idx >= 0:
                h = json_hallazgos[idx]
                json_status = h.get("human_review_status", "pending_validation")
                json_desc   = h.get("descripcion", "")[:200]
                json_notes  = h.get("human_review_notes", "") or ""

            entry = {
                "card_id":    card_id,
                "card_name":  card_info["card_name"],
                "vistas":     card_info["vistas"],
                "patron":     r["patron"],
                "desc_excel": r["desc"][:200],
                "json_status": json_status,
                "json_desc":   json_desc,
                "json_notes":  json_notes[:200],
                "match_method": method,
                "match_score":  round(score, 3),
            }

            if json_status in DECIDED_STATUSES:
                grupo_b_json_decidido.append(entry)
                if verbose and (grupo in ("all", "b")):
                    print(f"\n  [GRUPO B — AUTO-SYNC] card #{card_id} ({card_info['vistas']:,}v)")
                    print(f"    JSON decidió: {json_status}")
                    print(f"    Notas: {json_notes[:80]}")
            else:
                grupo_b_json_pending.append(entry)
                if verbose and (grupo in ("all", "b")):
                    print(f"\n  [GRUPO B — PENDING] card #{card_id} ({card_info['vistas']:,}v)")
                    print(f"    Patrón: {r['patron']}")
                    print(f"    Desc:   {r['desc'][:100]}")

    # ─────────────────────────────────────────────
    # Estadísticas grupo A por causa
    # ─────────────────────────────────────────────
    causas_a = defaultdict(int)
    for e in grupo_a:
        causas_a[e["causa"]] += 1

    report = {
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "resumen": {
            "grupo_a_sin_match":           len(grupo_a),
            "grupo_a_reparables_fuzzy":    sum(1 for e in grupo_a if e["causa"].startswith("fuzzy")),
            "grupo_a_force_match":         causas_a.get("single_hallazgo_force_match_possible", 0),
            "grupo_a_sin_json":            causas_a.get("card_not_in_json", 0),
            "grupo_b_total_pending":       len(pending_rows),
            "grupo_b_auto_sync":           len(grupo_b_json_decidido),
            "grupo_b_json_pending":        len(grupo_b_json_pending),
            "grupo_b_sin_json":            len(grupo_b_sin_json),
        },
        "causas_grupo_a": dict(causas_a),
        "grupo_a": grupo_a,
        "grupo_b_auto_sync":   grupo_b_json_decidido,
        "grupo_b_json_pending": grupo_b_json_pending,
        "grupo_b_sin_json":    grupo_b_sin_json,
    }
    return report


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Diagnóstico de pending Excel: sin match y pending con match"
    )
    parser.add_argument("--verbose", action="store_true", help="Muestra cada item en consola")
    parser.add_argument("--grupo",   default="all",
                        choices=["all", "a", "b"],
                        help="Filtra grupo a mostrar")
    parser.add_argument("--stats",   action="store_true", help="Solo estadísticas, sin guardar reporte")
    args = parser.parse_args()

    print("=" * 65)
    print("  Diagnose Pending Excel — Sin match + Pending con match")
    print("=" * 65)
    print()

    report = analyze(verbose=args.verbose, grupo=args.grupo)
    r = report["resumen"]

    print(f"\n  {'─'*55}")
    print(f"  GRUPO A — filas decididas en Excel sin match JSON")
    print(f"  {'─'*55}")
    print(f"    Total sin match:             {r['grupo_a_sin_match']:>4}")
    print(f"    ↳ Force-match posible (1h):  {r['grupo_a_force_match']:>4}  (card con 1 hallazgo)")
    print(f"    ↳ Reparable fuzzy (≥0.60):   {r['grupo_a_reparables_fuzzy']:>4}")
    print(f"    ↳ Card no existe en JSONs:   {r['grupo_a_sin_json']:>4}")
    print(f"    ↳ Genuinamente sin match:    {r['grupo_a_sin_match'] - r['grupo_a_force_match'] - r['grupo_a_reparables_fuzzy'] - r['grupo_a_sin_json']:>4}")

    if report.get("causas_grupo_a"):
        print(f"\n    Detalle por causa:")
        for causa, cnt in sorted(report["causas_grupo_a"].items(), key=lambda x: -x[1]):
            print(f"      {causa:<45} {cnt:>3}")

    print(f"\n  {'─'*55}")
    print(f"  GRUPO B — filas pending en Excel")
    print(f"  {'─'*55}")
    print(f"    Total pending en Excel:      {r['grupo_b_total_pending']:>4}")
    print(f"    ↳ JSON ya tiene decisión:    {r['grupo_b_auto_sync']:>4}  → auto-sync posible")
    print(f"    ↳ JSON también pending:      {r['grupo_b_json_pending']:>4}  → necesita SQL/data_owner")
    print(f"    ↳ Card no existe en JSONs:   {r['grupo_b_sin_json']:>4}")

    if r["grupo_b_auto_sync"] > 0:
        print(f"\n  ⚡ {r['grupo_b_auto_sync']} filas Excel pueden actualizarse desde el JSON sin revisión.")
        print(f"     Ejecuta: python src/core/propagate_reviews.py (actualiza Excel desde JSONs decididos)")

    if not args.stats:
        with open(REPORT_PATH, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2, default=str)
        print(f"\n  💾 Reporte guardado en: {REPORT_PATH.name}")
    else:
        print("\n  Modo --stats: no se guardó reporte.")

    print("=" * 65)

    # Acciones sugeridas
    total_reparable = r["grupo_a_force_match"] + r["grupo_a_reparables_fuzzy"]
    if total_reparable > 0:
        print(f"\n  Siguiente paso para Grupo A ({total_reparable} reparables):")
        print(f"    → python src/core/propagate_reviews.py  (con matching mejorado)")
    print()


if __name__ == "__main__":
    main()
