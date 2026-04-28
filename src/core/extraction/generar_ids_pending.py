"""
generar_ids_pending.py
======================
Lee el tracker Excel y genera ids_pending_508.txt con todos los card_id
cuyo human_review_status == 'pending_validation' (o vacío).

Correr desde Windows:
    python -X utf8 generar_ids_pending.py

Output:
    ids_pending_508.txt  — un ID por línea, listo para:
    python -X utf8 audit_agent.py --reaudit-file ids_pending_508.txt
"""

import openpyxl
from pathlib import Path

_ROOT   = Path(__file__).resolve().parent.parent.parent.parent
TRACKER = _ROOT / "data" / "raw" / "tracker_auditoria_metabase.xlsx"
OUTPUT  = _ROOT / "data" / "processed" / "resultados" / "ids_pending.txt"

# Hoja que contiene todos los hallazgos
TARGET_SHEET = "🔍 Todos los hallazgos"

wb = openpyxl.load_workbook(TRACKER, read_only=True, data_only=True)
print(f"Hojas disponibles: {wb.sheetnames}")

if TARGET_SHEET not in wb.sheetnames:
    print(f"\nERROR: No se encontró la hoja '{TARGET_SHEET}'")
    print("Probando con la primera hoja que tenga 'hallazgo' o 'finding' en el nombre...")
    TARGET_SHEET = next(
        (s for s in wb.sheetnames if "hallazgo" in s.lower() or "finding" in s.lower()),
        wb.sheetnames[-1]  # fallback: última hoja
    )
    print(f"Usando: '{TARGET_SHEET}'")

ws = wb[TARGET_SHEET]

# Buscar la fila de headers — puede ser fila 1 o 2 (a veces hay una fila de título decorativa)
headers = []
header_row_num = 1
for row_num, row in enumerate(ws.iter_rows(max_row=3), start=1):
    vals = [str(c.value).strip() if c.value is not None else "" for c in row]
    # La fila de headers tiene varias celdas llenas, no solo una
    non_empty = sum(1 for v in vals if v)
    if non_empty >= 3:
        headers = vals
        header_row_num = row_num
        break

print(f"\nHeaders encontrados en fila {header_row_num} de '{TARGET_SHEET}':")
for i, h in enumerate(headers):
    if h:  # solo mostrar las no vacías
        print(f"  [{i}] {repr(h)}")

# Detectar columnas de card_id y human_review_status
COL_CARD_ID = None
COL_STATUS  = None

for i, h in enumerate(headers):
    h_lower = h.lower().strip()
    if h_lower in ("card id", "card_id", "id"):
        COL_CARD_ID = i
    if "validation status" in h_lower or "human_review_status" in h_lower or "review_status" in h_lower:
        COL_STATUS = i

if COL_CARD_ID is None:
    print("\nERROR: No se encontró columna de card_id.")
    print("Edita el script y añade el índice correcto manualmente:")
    print("  COL_CARD_ID = <número de la columna de arriba>")
    print("  COL_STATUS  = <número de la columna de status>")
    exit(1)

print(f"\ncard_id  → col {COL_CARD_ID} ('{headers[COL_CARD_ID]}')")
print(f"status   → col {COL_STATUS}  ('{headers[COL_STATUS] if COL_STATUS is not None else 'NO ENCONTRADA'}')")

if COL_STATUS is None:
    print("\nERROR: No se encontró columna de human_review_status.")
    exit(1)

# --- Extraer IDs pending ---
pending_ids = set()
total = 0

for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
    total += 1
    card_id = row[COL_CARD_ID] if COL_CARD_ID < len(row) else None
    status  = row[COL_STATUS]  if COL_STATUS  < len(row) else None

    if card_id is None:
        continue

    status_str = str(status).strip().lower() if status else ""
    # El tracker Excel usa "Validation Status" con valores como:
    # pending_validation, pending, None/vacío = pendiente
    # false_positive, confirmed_error, intentional = ya revisado
    if status_str in ("pending_validation", "pending", "none", "", "null"):
        try:
            pending_ids.add(int(card_id))
        except (ValueError, TypeError):
            pass

ids_sorted = sorted(pending_ids)

with open(OUTPUT, "w", encoding="utf-8") as f:
    for cid in ids_sorted:
        f.write(f"{cid}\n")

print(f"\nTotal filas procesadas : {total}")
print(f"IDs pending encontrados: {len(ids_sorted)}")
print(f"Archivo generado       : {OUTPUT}")
print(f"\nPróximo paso:")
print(f"  python -X utf8 audit_agent.py --reaudit-file {OUTPUT}")
